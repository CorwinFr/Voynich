import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# === 1. Load all corpora ===

# Antidotarium Nicolai
an = pd.read_csv('/sessions/laughing-jolly-bell/mnt/Voynich FINAL/Clean/antidotarium_nicolai_ingredients.csv')
an = an.rename(columns={'Ingredient_Latin': 'Latin'})

# Circa Instans
ci = pd.read_csv('/sessions/laughing-jolly-bell/mnt/Voynich FINAL/Clean/circa_instans_galenic_degrees.csv')

# Lylye of Medicynes
lylye = pd.read_csv('/sessions/laughing-jolly-bell/mnt/Voynich FINAL/Clean/lylye_medicynes_ingredients.csv')

# V39 decoded substances (from previous analysis)
v39_substances = {
    'Rosa': 89, 'Aqua': 847, 'Oleum': 312, 'Acetum': 156, 'Sal': 98,
    'Cera': 76, 'Mel': 67, 'Succus': 54, 'Opium': 45, 'Crocus': 38,
    'Myrrha': 34, 'Cinnamomum': 28, 'Aloe': 24, 'Piper': 21,
    'Mastix': 18, 'Camphora': 15, 'Glycyrrhiza': 12, 'Zingiber': 11,
    'Absinthium': 9, 'Ruta': 8, 'Viola': 7, 'Plantago': 6,
    'Hyssopus': 5, 'Salvia': 5, 'Origanum': 4, 'Mentha': 4,
    'Foeniculum': 3, 'Petroselinum': 3, 'Gentiana': 2, 'Lavandula': 2,
    'Mandragora': 2, 'Papaver': 1
}

# === 2. Build master Latin ingredient list ===

# Collect all Latin names from AN
all_latin = set(an['Latin'].dropna().unique())

# Add from CI
ci_latin_map = {}
for _, row in ci.iterrows():
    latin = str(row.get('Latin', '')).strip()
    if latin and latin != 'nan':
        all_latin.add(latin)
        ci_latin_map[latin] = row

# Add from Lylye (Latin equivalents)
lylye_latin_map = {}
for _, row in lylye.iterrows():
    latin = str(row.get('Latin_Equivalent', '')).strip()
    if latin and latin != 'nan' and not latin.startswith('Aqua_') and not latin.startswith('Mel_') and not latin.startswith('Oleum_') and not latin.startswith('Albumen') and not latin.startswith('Lac_'):
        all_latin.add(latin)
        if latin not in lylye_latin_map:
            lylye_latin_map[latin] = row

# Add V39
for lat in v39_substances:
    all_latin.add(lat)

# Remove compound/prepared forms
remove = {'Aqua_rosarum', 'Mel_rosatum', 'Oleum_rosatum', 'Albumen_ovi', 'Lac_mulieris', 'Nux_moschata', 'Zuccarum', 'Sal'}
all_latin -= remove

# === 3. Build cross-reference rows ===
rows = []

for latin in sorted(all_latin):
    row = {'Latin': latin}

    # AN data
    an_match = an[an['Latin'] == latin]
    if len(an_match) > 0:
        r = an_match.iloc[0]
        row['AN_Rang'] = int(r['Rang'])
        row['AN_Nb_Recettes'] = int(r['Nb_Recettes'])
        row['AN_Pct'] = float(r['Pct_Recettes'])
    else:
        row['AN_Rang'] = None
        row['AN_Nb_Recettes'] = None
        row['AN_Pct'] = None

    # CI data
    ci_match = ci[ci['Latin'].astype(str).str.strip() == latin]
    if len(ci_match) > 0:
        r = ci_match.iloc[0]
        row['CI_OldFrench'] = str(r.get('OldFrench', ''))
        row['CI_Thermal'] = str(r.get('Thermal', ''))
        row['CI_ThermalDeg'] = str(r.get('ThermalDeg', ''))
        row['CI_Moisture'] = str(r.get('Moisture', ''))
        row['CI_MoistureDeg'] = str(r.get('MoistureDeg', ''))
    else:
        row['CI_OldFrench'] = None
        row['CI_Thermal'] = None
        row['CI_ThermalDeg'] = None
        row['CI_Moisture'] = None
        row['CI_MoistureDeg'] = None

    # Lylye data
    lylye_match = lylye[lylye['Latin_Equivalent'].astype(str).str.strip() == latin]
    if len(lylye_match) > 0:
        r = lylye_match.iloc[0]
        row['Lylye_ME'] = str(r['Middle_English'])
        row['Lylye_Rang'] = int(r['Rang'])
        row['Lylye_Nb'] = int(r['Nb_Recettes'])
        row['Lylye_Pct'] = float(r['Pct_Recettes'])
    else:
        row['Lylye_ME'] = None
        row['Lylye_Rang'] = None
        row['Lylye_Nb'] = None
        row['Lylye_Pct'] = None

    # V39 data
    if latin in v39_substances:
        row['V39_Present'] = 'OUI'
        row['V39_Occurrences'] = v39_substances[latin]
    else:
        row['V39_Present'] = 'NON'
        row['V39_Occurrences'] = 0

    # Corpus count
    count = 0
    if row['AN_Rang'] is not None: count += 1
    if row['CI_OldFrench'] is not None: count += 1
    if row['Lylye_Rang'] is not None: count += 1
    if row['V39_Present'] == 'OUI': count += 1
    row['Nb_Corpora'] = count

    rows.append(row)

df = pd.DataFrame(rows)

# Sort by number of corpora (desc), then AN rank
df = df.sort_values(['Nb_Corpora', 'AN_Rang'], ascending=[False, True])
df = df.reset_index(drop=True)

# === 4. Create Excel workbook ===
wb = Workbook()

# --- Sheet 1: Master Cross-Reference ---
ws1 = wb.active
ws1.title = "Matrice Croisee"

# Colors
header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
sub_fill_an = PatternFill('solid', fgColor='D6E4F0')
sub_fill_ci = PatternFill('solid', fgColor='E2EFDA')
sub_fill_ly = PatternFill('solid', fgColor='FFF2CC')
sub_fill_v39 = PatternFill('solid', fgColor='FCE4D6')
sub_fill_synth = PatternFill('solid', fgColor='D9D2E9')
data_font = Font(name='Arial', size=9)
bold_font = Font(name='Arial', size=9, bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers row 1 (merged groups)
ws1.merge_cells('A1:A2')
ws1['A1'] = 'Ingredient Latin'
ws1['A1'].fill = header_fill
ws1['A1'].font = header_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('B1:D1')
ws1['B1'] = 'ANTIDOTARIUM NICOLAI'
ws1['B1'].fill = header_fill
ws1['B1'].font = header_font
ws1['B1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('E1:I1')
ws1['E1'] = 'CIRCA INSTANS (Galenic)'
ws1['E1'].fill = header_fill
ws1['E1'].font = header_font
ws1['E1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('J1:M1')
ws1['J1'] = 'LYLYE OF MEDICYNES'
ws1['J1'].fill = header_fill
ws1['J1'].font = header_font
ws1['J1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('N1:O1')
ws1['N1'] = 'VOYNICH V39'
ws1['N1'].fill = header_fill
ws1['N1'].font = header_font
ws1['N1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('P1:P2')
ws1['P1'] = 'Nb Corpora'
ws1['P1'].fill = header_fill
ws1['P1'].font = header_font
ws1['P1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Sub-headers row 2
sub_headers = [
    ('B2', 'Rang', sub_fill_an), ('C2', 'Nb Rec.', sub_fill_an), ('D2', '% Rec.', sub_fill_an),
    ('E2', 'Ancien Fr.', sub_fill_ci), ('F2', 'Thermal', sub_fill_ci), ('G2', 'Deg.T', sub_fill_ci),
    ('H2', 'Moisture', sub_fill_ci), ('I2', 'Deg.M', sub_fill_ci),
    ('J2', 'Middle Eng.', sub_fill_ly), ('K2', 'Rang', sub_fill_ly), ('L2', 'Nb Rec.', sub_fill_ly), ('M2', '% Rec.', sub_fill_ly),
    ('N2', 'Present', sub_fill_v39), ('O2', 'Occur.', sub_fill_v39),
]
for cell_ref, text, fill in sub_headers:
    ws1[cell_ref] = text
    ws1[cell_ref].fill = fill
    ws1[cell_ref].font = Font(name='Arial', bold=True, size=9)
    ws1[cell_ref].alignment = Alignment(horizontal='center', wrap_text=True)

# Fill data
oui_fill = PatternFill('solid', fgColor='C6EFCE')
non_fill = PatternFill('solid', fgColor='F2F2F2')
four_corpus_fill = PatternFill('solid', fgColor='E2EFDA')

for idx, row in df.iterrows():
    r = idx + 3  # Excel row (1-indexed, after 2 header rows)

    ws1.cell(row=r, column=1, value=row['Latin']).font = bold_font

    # AN
    ws1.cell(row=r, column=2, value=row['AN_Rang']).font = data_font
    ws1.cell(row=r, column=3, value=row['AN_Nb_Recettes']).font = data_font
    if row['AN_Pct'] is not None:
        c = ws1.cell(row=r, column=4, value=row['AN_Pct']/100)
        c.font = data_font
        c.number_format = '0.0%'

    # CI
    for j, col_name in enumerate(['CI_OldFrench', 'CI_Thermal', 'CI_ThermalDeg', 'CI_Moisture', 'CI_MoistureDeg'], start=5):
        val = row[col_name]
        if val and str(val) != 'nan':
            ws1.cell(row=r, column=j, value=str(val)).font = data_font

    # Lylye
    ws1.cell(row=r, column=10, value=row['Lylye_ME']).font = data_font
    ws1.cell(row=r, column=11, value=row['Lylye_Rang']).font = data_font
    ws1.cell(row=r, column=12, value=row['Lylye_Nb']).font = data_font
    if row['Lylye_Pct'] is not None:
        c = ws1.cell(row=r, column=13, value=row['Lylye_Pct']/100)
        c.font = data_font
        c.number_format = '0.0%'

    # V39
    v39_cell = ws1.cell(row=r, column=14, value=row['V39_Present'])
    v39_cell.font = data_font
    v39_cell.alignment = Alignment(horizontal='center')
    if row['V39_Present'] == 'OUI':
        v39_cell.fill = oui_fill

    ws1.cell(row=r, column=15, value=row['V39_Occurrences']).font = data_font

    # Nb corpora
    nc = ws1.cell(row=r, column=16, value=row['Nb_Corpora'])
    nc.font = bold_font
    nc.alignment = Alignment(horizontal='center')
    if row['Nb_Corpora'] == 4:
        nc.fill = four_corpus_fill

    # Borders
    for c in range(1, 17):
        ws1.cell(row=r, column=c).border = thin_border

# Column widths
col_widths = [20, 6, 7, 7, 16, 10, 6, 10, 6, 16, 6, 7, 7, 8, 7, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws1.freeze_panes = 'B3'

# Auto-filter
ws1.auto_filter.ref = f'A2:P{len(df)+2}'

# --- Sheet 2: Top 50 concordance (sorted by AN rank) ---
ws2 = wb.create_sheet("Top 50 AN")

top50_an = df[df['AN_Rang'].notna()].sort_values('AN_Rang').head(50)

ws2['A1'] = 'Top 50 ingredients Antidotarium Nicolai - Concordance multi-corpus'
ws2['A1'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
ws2.merge_cells('A1:H1')

headers2 = ['Latin', 'AN Rang', 'AN %', 'CI Thermal', 'CI Deg', 'Lylye Rang', 'V39', 'Nb Corpora']
for j, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=j, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center')

for idx, (_, row) in enumerate(top50_an.iterrows()):
    r = idx + 4
    ws2.cell(row=r, column=1, value=row['Latin']).font = bold_font
    ws2.cell(row=r, column=2, value=row['AN_Rang']).font = data_font
    if row['AN_Pct'] is not None:
        c = ws2.cell(row=r, column=3, value=row['AN_Pct']/100)
        c.number_format = '0.0%'
        c.font = data_font

    thermal = str(row['CI_Thermal']) if row['CI_Thermal'] and str(row['CI_Thermal']) != 'nan' else ''
    deg = str(row['CI_ThermalDeg']) if row['CI_ThermalDeg'] and str(row['CI_ThermalDeg']) != 'nan' else ''
    ws2.cell(row=r, column=4, value=thermal).font = data_font
    ws2.cell(row=r, column=5, value=deg).font = data_font

    ws2.cell(row=r, column=6, value=row['Lylye_Rang']).font = data_font

    v = ws2.cell(row=r, column=7, value=row['V39_Present'])
    v.font = data_font
    if row['V39_Present'] == 'OUI':
        v.fill = oui_fill

    ws2.cell(row=r, column=8, value=row['Nb_Corpora']).font = bold_font
    ws2.cell(row=r, column=8).alignment = Alignment(horizontal='center')

    for c in range(1, 9):
        ws2.cell(row=r, column=c).border = thin_border

ws2.column_dimensions['A'].width = 20
for col in 'BCDEFGH':
    ws2.column_dimensions[col].width = 12

# --- Sheet 3: V39 substances with cross-ref ---
ws3 = wb.create_sheet("V39 Substances")

ws3['A1'] = 'Substances identifiees dans V39 decode - Concordance'
ws3['A1'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
ws3.merge_cells('A1:G1')

headers3 = ['Latin', 'V39 Occur.', 'AN Rang', 'AN %', 'CI Thermal/Deg', 'Lylye Rang', 'Commentaire']
for j, h in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=j, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center')

v39_df = df[df['V39_Present'] == 'OUI'].sort_values('V39_Occurrences', ascending=False)

for idx, (_, row) in enumerate(v39_df.iterrows()):
    r = idx + 4
    ws3.cell(row=r, column=1, value=row['Latin']).font = bold_font
    ws3.cell(row=r, column=2, value=row['V39_Occurrences']).font = data_font
    ws3.cell(row=r, column=3, value=row['AN_Rang']).font = data_font
    if row['AN_Pct'] is not None:
        c = ws3.cell(row=r, column=4, value=row['AN_Pct']/100)
        c.number_format = '0.0%'
        c.font = data_font

    thermal = str(row['CI_Thermal']) if row['CI_Thermal'] and str(row['CI_Thermal']) != 'nan' else ''
    deg = str(row['CI_ThermalDeg']) if row['CI_ThermalDeg'] and str(row['CI_ThermalDeg']) != 'nan' else ''
    ws3.cell(row=r, column=5, value=f"{thermal} {deg}".strip()).font = data_font
    ws3.cell(row=r, column=6, value=row['Lylye_Rang']).font = data_font

    # Comments
    comment = ''
    if row['AN_Rang'] is not None and row['AN_Rang'] <= 20:
        comment = 'TOP 20 AN'
    elif row['AN_Rang'] is not None and row['AN_Rang'] <= 50:
        comment = 'Top 50 AN'
    if row['Nb_Corpora'] == 4:
        comment += ' | 4 corpora' if comment else '4 corpora'
    ws3.cell(row=r, column=7, value=comment).font = data_font

    for c in range(1, 8):
        ws3.cell(row=r, column=c).border = thin_border

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['G'].width = 20
for col in 'BCDF':
    ws3.column_dimensions[col].width = 10

# --- Sheet 4: Statistics ---
ws4 = wb.create_sheet("Statistiques")

ws4['A1'] = 'Statistiques de couverture inter-corpus'
ws4['A1'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
ws4.merge_cells('A1:D1')

stats = [
    ('', '', '', ''),
    ('Corpus', 'Nb ingredients', 'Nb recettes', 'Source'),
    ('Antidotarium Nicolai', 114, 137, 'Van den Berg 1917 (editio princeps 1471)'),
    ('Circa Instans', 207, '-', 'Dorveaux 1913 (MS 3113 Ste-Genevieve)'),
    ('Lylye of Medicynes', 715, 421, 'Ancientbiotics mBio 2020'),
    ('Voynich V39 (decode)', 32, '-', 'K&A v12 pipeline'),
    ('', '', '', ''),
    ('Couverture croisee', 'Nombre', '', ''),
    ('Ingredients dans 4 corpora', len(df[df['Nb_Corpora']==4]), '', ''),
    ('Ingredients dans 3 corpora', len(df[df['Nb_Corpora']==3]), '', ''),
    ('Ingredients dans 2 corpora', len(df[df['Nb_Corpora']==2]), '', ''),
    ('Ingredients dans 1 corpus', len(df[df['Nb_Corpora']==1]), '', ''),
    ('Total ingredients uniques', len(df), '', ''),
    ('', '', '', ''),
    ('AN ingredients avec CI', len(an[an['CI_Thermal'].notna() & (an['CI_Thermal'] != '')]), 'sur 114', ''),
    ('V39 substances dans AN', len(v39_df[v39_df['AN_Rang'].notna()]), 'sur 32', ''),
]

for i, (a, b, c, d) in enumerate(stats, 3):
    ws4.cell(row=i, column=1, value=a).font = bold_font if i in [4, 10] else data_font
    ws4.cell(row=i, column=2, value=b).font = data_font
    ws4.cell(row=i, column=3, value=c).font = data_font
    ws4.cell(row=i, column=4, value=d).font = data_font
    if i == 4 or i == 10:
        for col in range(1, 5):
            ws4.cell(row=i, column=col).fill = header_fill
            ws4.cell(row=i, column=col).font = header_font

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 42

# Save
output_path = '/sessions/laughing-jolly-bell/mnt/Voynich FINAL/Clean/matrice_croisee_4_corpora.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total unique ingredients: {len(df)}")
print(f"4 corpora: {len(df[df['Nb_Corpora']==4])}")
print(f"3 corpora: {len(df[df['Nb_Corpora']==3])}")
print(f"2 corpora: {len(df[df['Nb_Corpora']==2])}")
print(f"1 corpus: {len(df[df['Nb_Corpora']==1])}")
print(f"\nIngredients in ALL 4 corpora:")
four = df[df['Nb_Corpora']==4].sort_values('AN_Rang')
for _, r in four.iterrows():
    print(f"  {r['Latin']:20s} AN#{int(r['AN_Rang']):3d} ({r['AN_Pct']:.1f}%)  CI:{str(r['CI_Thermal'])[:3]} {str(r['CI_ThermalDeg'])}  Lylye#{int(r['Lylye_Rang']):3d}  V39:{r['V39_Occurrences']}")
